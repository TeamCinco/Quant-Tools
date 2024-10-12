import pandas as pd
import numpy as np
import yfinance as yf
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from tqdm import tqdm
from datetime import datetime

def get_stock_data(ticker):
    stock_data = yf.download(ticker, period='5y')[['Open', 'Close']]
    return stock_data

def monte_carlo_simulation(current_price, mean_return, std_return, days, mc_runs):
    simulations = np.zeros((days, mc_runs))
    for run in tqdm(range(mc_runs), desc="Running Monte Carlo Simulations"):
        daily_returns = np.random.normal(mean_return, std_return, days)
        price_paths = current_price * np.exp(np.cumsum(daily_returns))
        simulations[:, run] = price_paths
    return simulations

def calculate_aroc(simulations):
    aroc_data = []
    total_aroc = 0
    for i in range(simulations.shape[1]):
        start_price = simulations[0, i]
        end_price = simulations[-1, i]
        aroc = (end_price - start_price) / start_price * 100
        total_aroc += aroc
        log_aroc = np.log1p(aroc / 100)  # Log AROC
        sigma = np.random.randn()
        log_sigma = np.log1p(sigma)
        percent_change = (end_price - start_price) / start_price * 100
        
        sigma_range = 'Negative' if sigma < 0 else 'Neutral' if sigma <= 1 else 'Positive'
        
        aroc_data.append({
            'MC Path': i + 1,
            'Date': datetime.now().strftime('%Y-%m-%d'),
            'AROC': aroc,
            'Log AROC': log_aroc,
            'Sigma': sigma,
            'Log Sigma': log_sigma,
            'Sigma Range': sigma_range,
            'Percent Change': percent_change,
            'Start Price': start_price,
            'End Price': end_price
        })
    
    # Calculate the mean AROC across all paths
    mean_aroc = total_aroc / simulations.shape[1]
    
    # Add mean AROC to each path's data
    for data in aroc_data:
        data['Mean AROC'] = mean_aroc

    return aroc_data

def save_to_excel(aroc_data, filename):
    wb = Workbook()
    ws_neg = wb.active
    ws_neg.title = "Negative Sigma"
    ws_neutral = wb.create_sheet("Neutral Sigma")
    ws_pos = wb.create_sheet("Positive Sigma")

    headers = ['MC Path', 'Date', 'AROC', 'Log AROC', 'Sigma', 'Log Sigma', 'Sigma Range', 'Percent Change', 'Start Price', 'End Price', 'Mean AROC']
    ws_neg.append(headers)
    ws_neutral.append(headers)
    ws_pos.append(headers)

    sorted_data = sorted(aroc_data, key=lambda x: x['AROC'], reverse=True)

    neg_data = [row for row in sorted_data if row['Sigma'] < 0]
    neutral_data = [row for row in sorted_data if 0 <= row['Sigma'] <= 1]
    pos_data = [row for row in sorted_data if row['Sigma'] > 1]

    red_fill = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid")
    green_fill = PatternFill(start_color="FFCCFFCC", end_color="FFCCFFCC", fill_type="solid")

    for sheet, data_set in zip([ws_neg, ws_neutral, ws_pos], [neg_data, neutral_data, pos_data]):
        for idx, row in enumerate(data_set, start=2):
            sheet.append([row.get(header, 'N/A') for header in headers])
            percent_change = row.get('Percent Change', 0)
            fill = red_fill if percent_change < 0 else green_fill
            for col in range(1, len(headers) + 1):
                sheet.cell(row=idx, column=col).fill = fill

    wb.save(filename)
    print(f"Data saved to {filename}")

def plot_histogram(aroc_data, ticker):
    plt.figure(figsize=(10, 6))
    plt.hist([row['AROC'] for row in aroc_data], bins=30, color='blue', edgecolor='black', alpha=0.7)
    plt.title(f'Histogram of AROC for {ticker}')
    plt.xlabel('AROC (%)')
    plt.ylabel('Frequency')
    plt.grid(True)
    plt.show()

def main():
    ticker = input("Enter the stock ticker symbol: ").upper()
    days = int(input("Enter the number of days for the simulation: "))
    mc_runs = int(input("Enter the number of Monte Carlo runs: "))

    stock_data = get_stock_data(ticker)
    current_price = stock_data['Close'].iloc[-1]
    daily_returns = (stock_data['Close'] - stock_data['Open']) / stock_data['Open']
    mean_return = daily_returns.mean()
    std_return = daily_returns.std()

    simulations = monte_carlo_simulation(current_price, mean_return, std_return, days, mc_runs)
    aroc_data = calculate_aroc(simulations)

    excel_filename = f"{ticker}_MC_AROC.xlsx"
    save_to_excel(aroc_data, excel_filename)

    plot_histogram(aroc_data, ticker)

if __name__ == "__main__":
    main()
