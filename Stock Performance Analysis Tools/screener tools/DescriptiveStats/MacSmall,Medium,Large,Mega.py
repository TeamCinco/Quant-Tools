import json 
import yfinance as yf
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from scipy import stats
from concurrent.futures import ThreadPoolExecutor, as_completed
import random
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Color
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, ColorScale, FormatObject, ColorScaleRule

# Set default encoding to UTF-8 for console output
sys.stdout.reconfigure(encoding='utf-8')

def load_tickers(file_path):
    with open(file_path, 'r') as file:
        data = json.load(file)

    return {item['ticker']: item for item in data.values() if item['ticker'] != "N/A"}



def filter_by_market_cap(ticker, size):
    stock = yf.Ticker(ticker)
    info = stock.info
    market_cap = info.get('marketCap', None)

    if market_cap is None:
        print(f"Market cap data not available for {ticker}.")
        return False

    if size == 'small' and market_cap <= 2e9:
        return True
    elif size == 'medium' and 2e9 < market_cap <= 10e9:
        return True
    elif size == 'large' and 10e9 < market_cap < 200e9:
        return True
    elif size == 'mega' and market_cap >= 200e9:
        return True
    else:
        return False

    
def get_stock_data(ticker, days=30):
    stock = yf.Ticker(ticker)
    
    # Break down the total number of days into chunks that Yahoo Finance allows
    if days <= 5:
        period = '5d'
    elif days <= 30:
        period = '1mo'
    elif days <= 90:
        period = '3mo'
    elif days <= 180:
        period = '6mo'
    elif days <= 365:
        period = '1y'
    else:
        # For more than 1 year, we split the data into 6-month intervals
        end_date = pd.Timestamp.today()
        start_date = end_date - pd.DateOffset(days=days)
        hist = pd.DataFrame()

        while start_date < end_date:
            # Get data in chunks
            chunk_end = min(start_date + pd.DateOffset(months=6), end_date)
            chunk_hist = stock.history(start=start_date.strftime('%Y-%m-%d'),
                                       end=chunk_end.strftime('%Y-%m-%d'))
            hist = pd.concat([hist, chunk_hist])
            start_date = chunk_end

        return hist

    # If within allowed periods
    hist = stock.history(period=period)
    return hist


def calculate_aroc(hist):
    daily_returns = (hist['Open'] - hist['Close']) / hist['Open'] * 100
    aroc = daily_returns.mean()
    return aroc

def log_transform_aroc(aroc):
    return np.log(np.abs(aroc) + 1) * np.sign(aroc)

def process_ticker(ticker, data, size_filter, days):
    try:
        if not filter_by_market_cap(ticker, size_filter):
            print(f"{ticker} does not match the market cap size '{size_filter}'; skipping.")
            return None

        hist = get_stock_data(ticker, days)  # Pass the 'days' argument here
        if hist is None or hist.empty:
            print(f"No historical data for {ticker}; skipping.")
            return None

        aroc = calculate_aroc(hist)
        if not np.isfinite(aroc):
            print(f"Invalid AROC calculated for {ticker}; skipping.")
            return None

        log_aroc = log_transform_aroc(aroc)

        start_price = hist['Open'].iloc[0]
        end_price = hist['Close'].iloc[-1]
        percent_change = ((end_price - start_price) / start_price) * 100

        stock_data = {
            'Ticker': ticker,
            'CompanyName': data.get('CompanyName', ''),
            'CIK': data.get('CIK', ''),
            'SIC': data.get('SIC', ''),
            'SICDescription': data.get('SICDescription', ''),
            'OwnerOrg': data.get('OwnerOrg', 'N/A'),  # Added OwnerOrg column
            'Date': hist.index[-1].strftime('%Y-%m-%d'),
            'AROC': aroc,
            'Log AROC': log_aroc,
            'Start Price': start_price,
            'End Price': end_price,
            'Percent Change': percent_change
        }

        print(f"Processed {ticker} ({data.get('CompanyName', '')})")
        return stock_data, aroc, log_aroc, hist

    except Exception as e:
        if '404' in str(e):
            print(f"Error processing {ticker}: {e} (Data not found)")
        else:
            print(f"Error processing {ticker}: {e}")
        return None


    


def plot_trendline_and_std_ranges(hist, mean_aroc, std_deviation, mean_price, std_prices, ticker):
    plt.figure(figsize=(10, 6))
    x = np.arange(len(hist))
    y = hist['Close']

    # Plot the closing price
    plt.plot(hist.index, y, label='Closing Price', color='blue')

    # Calculate and plot the trendline
    slope, intercept, _, _, _ = stats.linregress(x, y)
    plt.plot(hist.index, slope * x + intercept, color='green', label='Trendline')

    # Repeat mean_price to match the length of hist.index
    mean_price_array = np.full(len(hist.index), mean_price)

    # Standard deviation lines using the provided std_prices
    plt.plot(hist.index, mean_price_array, color='orange', label='Mean')
    plt.plot(hist.index, mean_price_array + std_prices, color='yellow', linestyle='--', label='+1σ')
    plt.plot(hist.index, mean_price_array - std_prices, color='yellow', linestyle='--', label='-1σ')
    plt.plot(hist.index, mean_price_array + 2 * std_prices, color='red', linestyle='-.', label='+2σ')
    plt.plot(hist.index, mean_price_array - 2 * std_prices, color='red', linestyle='-.', label='-2σ')

    # Plot title and labels
    plt.title(f'{ticker} Stock Price with Trendline and Standard Deviation Ranges')
    plt.xlabel('Date')
    plt.ylabel('Price')
    plt.legend()
    plt.grid(True)

    # Save and display the plot
    plt.savefig(rf'C:\Users\cinco\Desktop\DATA FOR SCRIPTS\Results\9.30.24\{ticker}_Price_STD.png')
    print(f"Trendline and standard deviation chart '{ticker}_Price_STD.png' has been created.")
    plt.show()


def plot_performance_with_aroc_std(hist, mean_aroc, std_deviation, ticker):
    # Calculate daily performance (open - close) / open * 100
    daily_performance = (hist['Open'] - hist['Close']) / hist['Open'] * 100
    
    # Create the time series plot for daily performance
    plt.figure(figsize=(10, 6))
    plt.plot(hist.index, daily_performance, label='Daily Performance (%)', color='blue')

    # Plot the AROC mean and the standard deviation ranges
    plt.axhline(mean_aroc, color='orange', label='Mean AROC', linestyle='-')
    plt.axhline(mean_aroc + std_deviation, color='yellow', linestyle='--', label='+1σ')
    plt.axhline(mean_aroc - std_deviation, color='yellow', linestyle='--', label='-1σ')
    plt.axhline(mean_aroc + 2 * std_deviation, color='red', linestyle='-.', label='+2σ')
    plt.axhline(mean_aroc - 2 * std_deviation, color='red', linestyle='-.', label='-2σ')
    plt.axhline(mean_aroc + 3 * std_deviation, color='purple', linestyle=':', label='+3σ')
    plt.axhline(mean_aroc - 3 * std_deviation, color='purple', linestyle=':', label='-3σ')

    # Set title and labels
    plt.title(f'{ticker} Daily Performance with AROC and Standard Deviations')
    plt.xlabel('Date')
    plt.ylabel('Performance (%)')
    plt.legend()
    plt.grid(True)
    
    # Save and show the plot
    plt.savefig(f'/Users/jazzhashzzz/Desktop/data for scripts/results/screener/10.7.24/{ticker}_performance_aroc_std_chart.png')
    print(f"Performance chart with AROC and stds for {ticker} created.")
    plt.show()
    


def save_to_excel(data, filename):
    wb = Workbook()

    # Create sheets for each sigma range
    ws_neg = wb.active
    ws_neg.title = "Negative Sigma"
    ws_neutral = wb.create_sheet("Neutral Sigma")
    ws_pos = wb.create_sheet("Positive Sigma")

    # Write headers to each sheet
    headers = ['Ticker', 'SICDescription', 'OwnerOrg', 'CompanyName', 'CIK', 'SIC', 'Date', 'AROC', 'Log AROC', 'Sigma', 'Log Sigma', 'Sigma Range', 'Percent Change', 'Start Price', 'End Price', 'Mean AROC']
    ws_neg.append(headers)
    ws_neutral.append(headers)
    ws_pos.append(headers)

    # Sort data by AROC
    sorted_data = sorted(data, key=lambda x: x['AROC'], reverse=True)

    # Preprocess and split data based on sigma ranges
    neg_data = []
    neutral_data = []
    pos_data = []
    for row in sorted_data:
        if row['Sigma'] < 0:
            neg_data.append(row)
        elif 0 <= row['Sigma'] <= 1:
            neutral_data.append(row)
        else:
            pos_data.append(row)

    # Define fills for negative and positive percent changes
    red_fill = PatternFill(start_color="FFFFCCCC", end_color="FFFFCCCC", fill_type="solid")
    green_fill = PatternFill(start_color="FFCCFFCC", end_color="FFCCFFCC", fill_type="solid")

    # Write data to the respective sheets and apply row formatting based on percent change
    for sheet, data_set in zip([ws_neg, ws_neutral, ws_pos], [neg_data, neutral_data, pos_data]):
        for idx, row in enumerate(data_set, start=2):
            # Append row data
            sheet.append([row.get(header, 'N/A') for header in headers])

            # Apply red fill for negative percent change, green fill for positive percent change
            percent_change = row.get('Percent Change', 0)
            if percent_change < 0:
                fill = red_fill
            else:
                fill = green_fill

            # Apply fill to the entire row except the Sigma column
            for col in range(1, len(headers) + 1):  # len(headers) + 1 to cover all columns
                if col != 10:  # Skip 'Sigma' column which is the 10th column (index 9)
                    sheet.cell(row=idx, column=col).fill = fill

    # Apply conditional formatting to the Sigma column based on its values
    for sheet, data_set in zip([ws_neg, ws_neutral, ws_pos], [neg_data, neutral_data, pos_data]):
        if len(data_set) > 0:
            if sheet.title == "Negative Sigma":
                # Apply negative sigma gradient (dark red to light red)
                sigma_rule = ColorScaleRule(
                    start_type='min', start_color='FF0000',
                    mid_type='percentile', mid_value=50, mid_color='FF9999',
                    end_type='max', end_color='FFFFFF'
                )
            elif sheet.title == "Neutral Sigma":
                # Apply neutral sigma gradient (yellow to green)
                sigma_rule = ColorScaleRule(
                    start_type='min', start_color='FFFF00',
                    mid_type='percentile', mid_value=50, mid_color='FFFFFF',
                    end_type='max', end_color='00FF00'
                )
            else:
                # Apply positive sigma gradient (light green to dark green)
                sigma_rule = ColorScaleRule(
                    start_type='min', start_color='FFFFFF',
                    mid_type='percentile', mid_value=50, mid_color='99FF99',
                    end_type='max', end_color='00FF00'
                )

            # Apply the gradient to the 'Sigma' column (column J, index 10)
            sheet.conditional_formatting.add(f'J2:J{len(data_set) + 1}', sigma_rule)

    wb.save(filename)
    print(f"Data saved to {filename}")


# Now modify the main loop to include this function call


def main():
    tickers = load_tickers('/Users/jazzhashzzz/Desktop/data for scripts/tickers and rates/tickers.json')
    size_filter = input("Enter market cap size to filter by (small, medium, large): ").lower()
    
    # Ask the user for the number of days
    days = int(input("Enter the number of days for historical data: "))

    selected_tickers = list(tickers.items())

    stock_data_list = []
    aroc_list = []
    log_aroc_list = []

    # Pass the 'days' argument to the process_ticker function
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(process_ticker, ticker, cik, size_filter, days): ticker for ticker, cik in selected_tickers}

        for future in as_completed(futures):
            result = future.result()
            if result is not None:
                stock_data, aroc, log_aroc, _ = result
                stock_data_list.append(stock_data)
                aroc_list.append(aroc)
                log_aroc_list.append(log_aroc)

    if len(aroc_list) == 0:
        print("No valid AROC data available.")
        return

    # Use estimated population mean and variance
    mean_aroc = np.mean(aroc_list)
    population_variance = np.var(aroc_list, ddof=0)  # ddof=0 for population variance
    std_deviation = np.sqrt(population_variance)

    mean_log_aroc = np.mean(log_aroc_list)
    log_population_variance = np.var(log_aroc_list, ddof=0)
    std_deviation_log = np.sqrt(log_population_variance)

    # Prepare data for Excel
    excel_data = []
    for stock in stock_data_list:
        sigma = (stock['AROC'] - mean_aroc) / std_deviation
        log_sigma = (stock['Log AROC'] - mean_log_aroc) / std_deviation_log
        sigma_range = 'Within 1σ' if abs(sigma) <= 1 else ('Within 2σ' if abs(sigma) <= 2 else 'Beyond 2σ')
        excel_data.append({
            'Ticker': stock['Ticker'],
            'CompanyName': stock['CompanyName'],
            'CIK': stock['CIK'],
            'SIC': stock['SIC'],
            'SICDescription': stock['SICDescription'],
            'Date': stock['Date'],
            'AROC': stock['AROC'],
            'Log AROC': stock['Log AROC'],
            'Sigma': sigma,
            'Log Sigma': log_sigma,
            'Sigma Range': sigma_range,
            'Percent Change': stock['Percent Change'],
            'Start Price': stock['Start Price'],
            'End Price': stock['End Price'],
            'Mean AROC': mean_aroc
        })

    # Save to Excel
    excel_filename = f"/Users/jazzhashzzz/Desktop/data for scripts/results/screener/10.7.24/{size_filter}_Cap_Analysis90d.xlsx"
    save_to_excel(excel_data, excel_filename)
    
    def plot_aroc_bell_curve(aroc_list, log_aroc_list):
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(20, 6))
        
        # Original AROC distribution
        mean_aroc = np.mean(aroc_list)
        std_dev_aroc = np.std(aroc_list)
        
        x1 = np.linspace(mean_aroc - 4*std_dev_aroc, mean_aroc + 4*std_dev_aroc, 100)
        y1 = stats.norm.pdf(x1, mean_aroc, std_dev_aroc)
        
        ax1.plot(x1, y1, 'b-', label='Normal Distribution')
        ax1.hist(aroc_list, bins=50, density=True, alpha=0.7, color='g', label='AROC Distribution')
        
        ax1.set_title('Original AROC Distribution')
        ax1.set_xlabel('AROC (%)')
        ax1.set_ylabel('Density')
        ax1.legend()
        ax1.grid(True)
        
        # Log-transformed AROC distribution
        mean_log_aroc = np.mean(log_aroc_list)
        std_dev_log_aroc = np.std(log_aroc_list)
        
        x2 = np.linspace(mean_log_aroc - 4*std_dev_log_aroc, mean_log_aroc + 4*std_dev_log_aroc, 100)
        y2 = stats.norm.pdf(x2, mean_log_aroc, std_dev_log_aroc)
        
        ax2.plot(x2, y2, 'b-', label='Normal Distribution')
        ax2.hist(log_aroc_list, bins=50, density=True, alpha=0.7, color='r', label='Log AROC Distribution')
        
        ax2.set_title('Log-transformed AROC Distribution')
        ax2.set_xlabel('Log AROC')
        ax2.set_ylabel('Density')
        ax2.legend()
        ax2.grid(True)
        
        plt.tight_layout()
        output_dir = "/Users/jazzhashzzz/Desktop/data for scripts/results/screener/10.7.24"
        os.makedirs(output_dir, exist_ok=True)  # This will create the directory if it doesn't exist

        path = rf"{output_dir}/{size_filter}_aroc_bell_curves90d.png"
        plt.savefig(path)

        print("AROC bell curve charts 'aroc_bell_curves.png' have been created.")
        plt.show()
    
    # Plot AROC bell curves
    plot_aroc_bell_curve(aroc_list, log_aroc_list)

    while True:
        comparison_ticker = input("Enter a ticker to analyze (or 'quit' to exit): ").upper()
        if comparison_ticker.lower() == 'quit':
            break

        comparison_hist = get_stock_data(comparison_ticker)
        if comparison_hist is None:
            print(f"No data found for {comparison_ticker}")
            continue

        comparison_aroc = calculate_aroc(comparison_hist)
        comparison_sigma = (comparison_aroc - mean_aroc) / std_deviation

        print(f"{comparison_ticker} AROC: {comparison_aroc:.2f}%")
        print(f"{comparison_ticker} Sigma: {comparison_sigma:.2f}")

        mean_price = comparison_hist['Close'].mean()
        std_prices = comparison_hist['Close'].std()

        # Call the plot for the trendline and std ranges (existing plot)
        plot_trendline_and_std_ranges(comparison_hist, mean_aroc, std_deviation, mean_price, std_prices, comparison_ticker)

        # Call the new plot for daily performance vs AROC mean and stds
        plot_performance_with_aroc_std(comparison_hist, mean_aroc, std_deviation, comparison_ticker)



def main():
    tickers = load_tickers('/Users/jazzhashzzz/Desktop/data for scripts/tickers and rates/tickers.json')
    size_filter = input("Enter market cap size to filter by (small, medium, large, mega): ").lower()
    
    # Ask the user for the number of days
    days = int(input("Enter the number of days for historical data: "))

    selected_tickers = list(tickers.items())

    stock_data_list = []
    aroc_list = []
    log_aroc_list = []

    # Pass the 'days' argument to the process_ticker function
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(process_ticker, ticker, cik, size_filter, days): ticker for ticker, cik in selected_tickers}

        for future in as_completed(futures):
            result = future.result()
            if result is not None:
                stock_data, aroc, log_aroc, _ = result
                stock_data_list.append(stock_data)
                aroc_list.append(aroc)
                log_aroc_list.append(log_aroc)

    if len(aroc_list) == 0:
        print("No valid AROC data available.")
        return

    # Use estimated population mean and variance
    mean_aroc = np.mean(aroc_list)
    population_variance = np.var(aroc_list, ddof=0)  # ddof=0 for population variance
    std_deviation = np.sqrt(population_variance)

    mean_log_aroc = np.mean(log_aroc_list)
    log_population_variance = np.var(log_aroc_list, ddof=0)
    std_deviation_log = np.sqrt(log_population_variance)

    # Prepare data for Excel
    excel_data = []
    for stock in stock_data_list:
        sigma = (stock['AROC'] - mean_aroc) / std_deviation
        log_sigma = (stock['Log AROC'] - mean_log_aroc) / std_deviation_log
        sigma_range = 'Within 1σ' if abs(sigma) <= 1 else ('Within 2σ' if abs(sigma) <= 2 else 'Beyond 2σ')
        excel_data.append({
            'Ticker': stock['Ticker'],
            'CompanyName': stock['CompanyName'],
            'CIK': stock['CIK'],
            'SIC': stock['SIC'],
            'SICDescription': stock['SICDescription'],
            'Date': stock['Date'],
            'AROC': stock['AROC'],
            'Log AROC': stock['Log AROC'],
            'Sigma': sigma,
            'Log Sigma': log_sigma,
            'Sigma Range': sigma_range,
            'Percent Change': stock['Percent Change'],
            'Start Price': stock['Start Price'],
            'End Price': stock['End Price'],
            'Mean AROC': mean_aroc
        })

    # Save to Excel
    excel_filename = f"/Users/jazzhashzzz/Desktop/data for scripts/results/screener/10.7.24/{size_filter}_Cap_Analysis90d.xlsx"
    save_to_excel(excel_data, excel_filename)
    
    def plot_aroc_bell_curve(aroc_list, log_aroc_list):
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(20, 6))
        
        # Original AROC distribution
        mean_aroc = np.mean(aroc_list)
        std_dev_aroc = np.std(aroc_list)
        
        x1 = np.linspace(mean_aroc - 4*std_dev_aroc, mean_aroc + 4*std_dev_aroc, 100)
        y1 = stats.norm.pdf(x1, mean_aroc, std_dev_aroc)
        
        ax1.plot(x1, y1, 'b-', label='Normal Distribution')
        ax1.hist(aroc_list, bins=50, density=True, alpha=0.7, color='g', label='AROC Distribution')
        
        ax1.set_title('Original AROC Distribution')
        ax1.set_xlabel('AROC (%)')
        ax1.set_ylabel('Density')
        ax1.legend()
        ax1.grid(True)
        
        # Log-transformed AROC distribution
        mean_log_aroc = np.mean(log_aroc_list)
        std_dev_log_aroc = np.std(log_aroc_list)
        
        x2 = np.linspace(mean_log_aroc - 4*std_dev_log_aroc, mean_log_aroc + 4*std_dev_log_aroc, 100)
        y2 = stats.norm.pdf(x2, mean_log_aroc, std_dev_log_aroc)
        
        ax2.plot(x2, y2, 'b-', label='Normal Distribution')
        ax2.hist(log_aroc_list, bins=50, density=True, alpha=0.7, color='r', label='Log AROC Distribution')
        
        ax2.set_title('Log-transformed AROC Distribution')
        ax2.set_xlabel('Log AROC')
        ax2.set_ylabel('Density')
        ax2.legend()
        ax2.grid(True)
        
        plt.tight_layout()
        output_dir = "/Users/jazzhashzzz/Desktop/data for scripts/results/screener/10.7.24"
        os.makedirs(output_dir, exist_ok=True)  # This will create the directory if it doesn't exist

        path = rf"{output_dir}/{size_filter}_aroc_bell_curves90d.png"
        plt.savefig(path)

        print("AROC bell curve charts 'aroc_bell_curves.png' have been created.")
        plt.show()
    
    # Plot AROC bell curves
    plot_aroc_bell_curve(aroc_list, log_aroc_list)

    while True:
        comparison_ticker = input("Enter a ticker to analyze (or 'quit' to exit): ").upper()
        if comparison_ticker.lower() == 'quit':
            break

        comparison_hist = get_stock_data(comparison_ticker)
        if comparison_hist is None:
            print(f"No data found for {comparison_ticker}")
            continue

        comparison_aroc = calculate_aroc(comparison_hist)
        comparison_sigma = (comparison_aroc - mean_aroc) / std_deviation

        print(f"{comparison_ticker} AROC: {comparison_aroc:.2f}%")
        print(f"{comparison_ticker} Sigma: {comparison_sigma:.2f}")

        mean_price = comparison_hist['Close'].mean()
        std_prices = comparison_hist['Close'].std()

        # Call the plot for the trendline and std ranges (existing plot)
        plot_trendline_and_std_ranges(comparison_hist, mean_aroc, std_deviation, mean_price, std_prices, comparison_ticker)

        # Call the new plot for daily performance vs AROC mean and stds
        plot_performance_with_aroc_std(comparison_hist, mean_aroc, std_deviation, comparison_ticker)

if __name__ == "__main__":
    main()